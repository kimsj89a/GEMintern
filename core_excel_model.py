"""
Excel Financial Model Builder for GEM Intern.
Extracts investment structure from wiki + project documents,
then generates a PEF-style cash flow Excel model with openpyxl formulas.
"""
import io
import json
import logging
import re
from typing import Any, Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter

from ai_client import get_client
from google.genai import types

logger = logging.getLogger(__name__)


# ── Styles ──────────────────────────────────────────

_THIN_BORDER = Side(style="thin", color="CCCCCC")
_MEDIUM_BORDER = Side(style="medium", color="999999")

STYLES = {
    "title": {
        "font": Font(name="맑은 고딕", size=13, bold=True, color="FFFFFF"),
        "fill": PatternFill("solid", fgColor="1B2631"),
        "alignment": Alignment(horizontal="left", vertical="center"),
    },
    "header": {
        "font": Font(name="맑은 고딕", size=9, bold=True, color="FFFFFF"),
        "fill": PatternFill("solid", fgColor="2C3E50"),
        "alignment": Alignment(horizontal="center", vertical="center"),
        "border": Border(bottom=_MEDIUM_BORDER),
    },
    "section": {
        "font": Font(name="맑은 고딕", size=9, bold=True),
        "fill": PatternFill("solid", fgColor="D5D8DC"),
        "alignment": Alignment(horizontal="left", vertical="center"),
    },
    "input": {
        "font": Font(name="맑은 고딕", size=9, bold=True, color="0000CC"),
        "fill": PatternFill("solid", fgColor="FFF2CC"),
        "alignment": Alignment(horizontal="right", vertical="center"),
        "border": Border(
            top=_THIN_BORDER, bottom=_THIN_BORDER,
            left=_THIN_BORDER, right=_THIN_BORDER,
        ),
    },
    "data": {
        "font": Font(name="맑은 고딕", size=9),
        "alignment": Alignment(horizontal="right", vertical="center"),
        "border": Border(
            top=_THIN_BORDER, bottom=_THIN_BORDER,
            left=_THIN_BORDER, right=_THIN_BORDER,
        ),
    },
    "formula": {
        "font": Font(name="맑은 고딕", size=9),
        "alignment": Alignment(horizontal="right", vertical="center"),
        "border": Border(
            top=_THIN_BORDER, bottom=_THIN_BORDER,
            left=_THIN_BORDER, right=_THIN_BORDER,
        ),
    },
    "subtotal": {
        "font": Font(name="맑은 고딕", size=9, bold=True),
        "fill": PatternFill("solid", fgColor="F2F3F4"),
        "alignment": Alignment(horizontal="right", vertical="center"),
        "border": Border(bottom=_MEDIUM_BORDER),
    },
    "irr": {
        "font": Font(name="맑은 고딕", size=10, bold=True, color="0000CC"),
        "fill": PatternFill("solid", fgColor="E8F8F5"),
        "alignment": Alignment(horizontal="right", vertical="center"),
        "border": Border(
            top=_THIN_BORDER, bottom=_MEDIUM_BORDER,
            left=_THIN_BORDER, right=_THIN_BORDER,
        ),
    },
    "label": {
        "font": Font(name="맑은 고딕", size=9),
        "alignment": Alignment(horizontal="left", vertical="center", indent=1),
        "border": Border(
            top=_THIN_BORDER, bottom=_THIN_BORDER,
            left=_THIN_BORDER, right=_THIN_BORDER,
        ),
    },
    "label_bold": {
        "font": Font(name="맑은 고딕", size=9, bold=True),
        "alignment": Alignment(horizontal="left", vertical="center"),
        "border": Border(
            top=_THIN_BORDER, bottom=_THIN_BORDER,
            left=_THIN_BORDER, right=_THIN_BORDER,
        ),
    },
}

NUM_FMT_AMOUNT = '#,##0.0;[Red](#,##0.0);"-"'
NUM_FMT_PCT = '0.0%'
NUM_FMT_INT = '#,##0'


def _apply_style(cell, style_name: str, number_format: str | None = None):
    """Apply a named style to a cell."""
    s = STYLES[style_name]
    for attr in ("font", "fill", "alignment", "border"):
        if attr in s:
            setattr(cell, attr, s[attr])
    if number_format:
        cell.number_format = number_format


def _safe_float(v, default=0.0) -> float:
    if v is None:
        return default
    try:
        return float(v)
    except (ValueError, TypeError):
        return default


def _safe_int(v, default=0) -> int:
    if v is None:
        return default
    try:
        return int(v)
    except (ValueError, TypeError):
        return default


# ── AI Extraction ──────────────────────────────────────

def _parse_json_object(text: str) -> dict:
    """Extract JSON object from AI response with fallback strategies."""
    # Try direct parse
    text = text.strip()
    if text.startswith("{"):
        try:
            return json.loads(text)
        except json.JSONDecodeError:
            pass

    # Try code block extraction
    m = re.search(r"```(?:json)?\s*(\{[\s\S]*?\})\s*```", text)
    if m:
        try:
            return json.loads(m.group(1))
        except json.JSONDecodeError:
            pass

    # Try finding the outermost { ... }
    start = text.find("{")
    if start >= 0:
        depth = 0
        for i in range(start, len(text)):
            if text[i] == "{":
                depth += 1
            elif text[i] == "}":
                depth -= 1
                if depth == 0:
                    try:
                        return json.loads(text[start : i + 1])
                    except json.JSONDecodeError:
                        break
    return {}


def extract_deal_structure(
    api_key: str,
    model_name: str,
    project_name: str,
    owner_id: int | None = None,
) -> dict:
    """Extract investment deal structure from wiki + project documents via AI."""
    from core_wiki import load_wiki
    from core_rag import load_project_docs_dict
    from prompts import EXCEL_MODEL_PROMPTS

    # Load wiki
    wiki = load_wiki(project_name, owner_id=owner_id)
    wiki_text = ""
    if wiki and wiki.get("sections"):
        parts = []
        for sec in wiki["sections"]:
            parts.append(f"## {sec.get('title', '')}\n{sec.get('content', '')}")
        wiki_text = "\n\n".join(parts)

    # Load source docs (budget-limited)
    docs = load_project_docs_dict(project_name, owner_id=owner_id)
    doc_budget = 150_000
    doc_text_parts = []
    per_doc = max(3000, doc_budget // max(len(docs), 1))
    for fname, content in docs.items():
        truncated = content[:per_doc]
        doc_text_parts.append(f"### 📄 {fname}\n{truncated}")
    doc_text = "\n\n".join(doc_text_parts)

    # Build prompt
    system_prompt = EXCEL_MODEL_PROMPTS["extract_structure"]
    user_content = f"[위키 내용]\n{wiki_text}\n\n[소스 문서]\n{doc_text}"

    client = get_client(api_key)
    response = client.models.generate_content(
        model=model_name,
        contents=user_content,
        config=types.GenerateContentConfig(
            system_instruction=system_prompt,
            temperature=0.2,
            max_output_tokens=8000,
            response_mime_type="application/json",
        ),
    )

    raw = response.text if hasattr(response, "text") else str(response)
    structure = _parse_json_object(raw)

    if not structure:
        raise ValueError("AI가 투자구조를 추출하지 못했습니다. 위키 또는 소스 문서에 투자 관련 내용이 필요합니다.")

    return structure


# ── Excel Builder ──────────────────────────────────────

def build_excel_model(structure: dict) -> bytes:
    """Build a PEF cash flow Excel model from extracted structure. Returns xlsx bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Cash Flow Model"

    company = structure.get("company_name", "대상회사")
    investor = structure.get("investor_name", "투자자")
    total_amount = _safe_float(structure.get("total_amount", 0))
    max_maturity = _safe_int(structure.get("max_maturity_years", 5))
    tranches = structure.get("tranches") or []
    irr_guarantee = structure.get("irr_guarantee")
    exit_scenarios = structure.get("exit_scenarios") or []
    covenants = structure.get("covenants") or []
    projections = structure.get("projections")

    if max_maturity < 1:
        max_maturity = 5
    if not tranches:
        tranches = [{"name": "Tranche A", "type": "CB", "amount": total_amount or 100,
                      "coupon_rate": 0, "ytm_base": 6, "ytm_stepup_per_year": 0,
                      "maturity_years": max_maturity}]

    n_years = max_maturity
    year_cols = list(range(n_years + 1))  # Year 0 ~ Year N

    # Column setup: A=Label, B=Unit, C~=Year0, Year1, ...
    label_col = 1  # A
    unit_col = 2   # B
    first_year_col = 3  # C = Year 0

    def yr_col(y: int) -> int:
        return first_year_col + y

    def yr_cell(row: int, y: int) -> str:
        return f"{get_column_letter(yr_col(y))}{row}"

    def cell_ref(row: int, col: int) -> str:
        return f"{get_column_letter(col)}{row}"

    # Column widths
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 8
    for y in year_cols:
        ws.column_dimensions[get_column_letter(yr_col(y))].width = 16

    row = 1

    # ── Title ──
    ws.merge_cells(start_row=row, start_column=1, end_row=row,
                   end_column=yr_col(n_years))
    title_cell = ws.cell(row=row, column=1,
                         value=f"{company} — PEF Cash Flow Model")
    _apply_style(title_cell, "title")
    ws.row_dimensions[row].height = 32
    row += 1

    # Subtitle
    ws.cell(row=row, column=1,
            value=f"투자자: {investor} | 총 투자규모: {total_amount:.0f}억원 | 만기: {max_maturity}년")
    ws.cell(row=row, column=1).font = Font(name="맑은 고딕", size=9, color="666666")
    row += 2

    # ── Year Headers ──
    header_row = row
    ws.cell(row=row, column=label_col, value="구분")
    _apply_style(ws.cell(row=row, column=label_col), "header")
    ws.cell(row=row, column=unit_col, value="단위")
    _apply_style(ws.cell(row=row, column=unit_col), "header")
    for y in year_cols:
        c = ws.cell(row=row, column=yr_col(y),
                    value=f"Year {y}" if y > 0 else "투자시점")
        _apply_style(c, "header")
    row += 1

    # ══════════════════════════════════════════════════
    # SECTION: Inputs
    # ══════════════════════════════════════════════════
    ws.cell(row=row, column=label_col, value="▶ Inputs")
    _apply_style(ws.cell(row=row, column=label_col), "section")
    for y in year_cols:
        _apply_style(ws.cell(row=row, column=yr_col(y)), "section")
    _apply_style(ws.cell(row=row, column=unit_col), "section")
    row += 1

    # Track input cell locations for formula references
    input_rows: Dict[str, int] = {}

    for ti, tr in enumerate(tranches):
        tname = tr.get("name", f"Tranche {chr(65 + ti)}")
        ttype = tr.get("type", "CB")
        amount = _safe_float(tr.get("amount", 0))
        coupon = _safe_float(tr.get("coupon_rate", 0))
        ytm_base = _safe_float(tr.get("ytm_base", 0))
        ytm_stepup = _safe_float(tr.get("ytm_stepup_per_year", 0))
        maturity = _safe_int(tr.get("maturity_years", max_maturity))
        div_rate = _safe_float(tr.get("dividend_rate", 0))
        call_year = _safe_int(tr.get("call_year", 0))
        put_year = _safe_int(tr.get("put_year", 0))

        # Tranche label
        ws.cell(row=row, column=label_col, value=f"  {tname} ({ttype})")
        _apply_style(ws.cell(row=row, column=label_col), "label_bold")
        row += 1

        # Amount
        key = f"t{ti}_amount"
        input_rows[key] = row
        ws.cell(row=row, column=label_col, value=f"    투자금액")
        _apply_style(ws.cell(row=row, column=label_col), "label")
        ws.cell(row=row, column=unit_col, value="억원")
        _apply_style(ws.cell(row=row, column=unit_col), "data")
        c = ws.cell(row=row, column=yr_col(0), value=amount)
        _apply_style(c, "input", NUM_FMT_AMOUNT)
        row += 1

        # Coupon rate
        key = f"t{ti}_coupon"
        input_rows[key] = row
        ws.cell(row=row, column=label_col, value=f"    이자율(Coupon)")
        _apply_style(ws.cell(row=row, column=label_col), "label")
        ws.cell(row=row, column=unit_col, value="%")
        _apply_style(ws.cell(row=row, column=unit_col), "data")
        c = ws.cell(row=row, column=yr_col(0), value=coupon / 100 if coupon else 0)
        _apply_style(c, "input", NUM_FMT_PCT)
        row += 1

        # YTM Base
        key = f"t{ti}_ytm_base"
        input_rows[key] = row
        ws.cell(row=row, column=label_col, value=f"    YTM (Base)")
        _apply_style(ws.cell(row=row, column=label_col), "label")
        ws.cell(row=row, column=unit_col, value="%")
        _apply_style(ws.cell(row=row, column=unit_col), "data")
        c = ws.cell(row=row, column=yr_col(0), value=ytm_base / 100 if ytm_base else 0)
        _apply_style(c, "input", NUM_FMT_PCT)
        row += 1

        # YTM Step-up
        key = f"t{ti}_ytm_stepup"
        input_rows[key] = row
        ws.cell(row=row, column=label_col, value=f"    YTM Step-up (/yr)")
        _apply_style(ws.cell(row=row, column=label_col), "label")
        ws.cell(row=row, column=unit_col, value="%")
        _apply_style(ws.cell(row=row, column=unit_col), "data")
        c = ws.cell(row=row, column=yr_col(0), value=ytm_stepup / 100 if ytm_stepup else 0)
        _apply_style(c, "input", NUM_FMT_PCT)
        row += 1

        # Maturity
        key = f"t{ti}_maturity"
        input_rows[key] = row
        ws.cell(row=row, column=label_col, value=f"    만기")
        _apply_style(ws.cell(row=row, column=label_col), "label")
        ws.cell(row=row, column=unit_col, value="년")
        _apply_style(ws.cell(row=row, column=unit_col), "data")
        c = ws.cell(row=row, column=yr_col(0), value=maturity)
        _apply_style(c, "input")
        row += 1

        # Dividend rate (if CPS)
        if ttype in ("CPS", "PS", "RCPS"):
            key = f"t{ti}_div_rate"
            input_rows[key] = row
            ws.cell(row=row, column=label_col, value=f"    우선배당률")
            _apply_style(ws.cell(row=row, column=label_col), "label")
            ws.cell(row=row, column=unit_col, value="%")
            _apply_style(ws.cell(row=row, column=unit_col), "data")
            c = ws.cell(row=row, column=yr_col(0), value=div_rate / 100 if div_rate else 0)
            _apply_style(c, "input", NUM_FMT_PCT)
            row += 1

        # Call/Put year
        if call_year:
            key = f"t{ti}_call_year"
            input_rows[key] = row
            ws.cell(row=row, column=label_col, value=f"    Call 행사시점")
            _apply_style(ws.cell(row=row, column=label_col), "label")
            ws.cell(row=row, column=unit_col, value="년")
            _apply_style(ws.cell(row=row, column=unit_col), "data")
            c = ws.cell(row=row, column=yr_col(0), value=call_year)
            _apply_style(c, "input")
            row += 1

        if put_year:
            key = f"t{ti}_put_year"
            input_rows[key] = row
            ws.cell(row=row, column=label_col, value=f"    Put 행사시점")
            _apply_style(ws.cell(row=row, column=label_col), "label")
            ws.cell(row=row, column=unit_col, value="년")
            _apply_style(ws.cell(row=row, column=unit_col), "data")
            c = ws.cell(row=row, column=yr_col(0), value=put_year)
            _apply_style(c, "input")
            row += 1

    # IRR Guarantee input
    if irr_guarantee and irr_guarantee.get("rate"):
        key = "irr_guarantee"
        input_rows[key] = row
        ws.cell(row=row, column=label_col, value="    IRR 보장률")
        _apply_style(ws.cell(row=row, column=label_col), "label")
        ws.cell(row=row, column=unit_col, value="%")
        _apply_style(ws.cell(row=row, column=unit_col), "data")
        c = ws.cell(row=row, column=yr_col(0),
                    value=_safe_float(irr_guarantee["rate"]) / 100)
        _apply_style(c, "input", NUM_FMT_PCT)
        row += 1

    row += 1  # blank

    # ══════════════════════════════════════════════════
    # SECTION: YTM/YTC Schedule (per tranche)
    # ══════════════════════════════════════════════════
    ws.cell(row=row, column=label_col, value="▶ YTM Schedule")
    _apply_style(ws.cell(row=row, column=label_col), "section")
    for y in year_cols:
        _apply_style(ws.cell(row=row, column=yr_col(y)), "section")
    _apply_style(ws.cell(row=row, column=unit_col), "section")
    row += 1

    ytm_schedule_rows: Dict[int, int] = {}  # ti -> row of YTM schedule
    ytm_accum_rows: Dict[int, int] = {}     # ti -> row of accumulated YTM factor

    for ti, tr in enumerate(tranches):
        tname = tr.get("name", f"Tranche {chr(65 + ti)}")
        ytm_base_row = input_rows.get(f"t{ti}_ytm_base")
        ytm_stepup_row = input_rows.get(f"t{ti}_ytm_stepup")

        # Effective YTM per year
        ytm_schedule_rows[ti] = row
        ws.cell(row=row, column=label_col, value=f"  {tname} 적용 YTM")
        _apply_style(ws.cell(row=row, column=label_col), "label")
        ws.cell(row=row, column=unit_col, value="%")
        _apply_style(ws.cell(row=row, column=unit_col), "data")

        if ytm_base_row and ytm_stepup_row:
            amt_cell_base = cell_ref(ytm_base_row, yr_col(0))
            amt_cell_stepup = cell_ref(ytm_stepup_row, yr_col(0))
            for y in year_cols:
                if y == 0:
                    ws.cell(row=row, column=yr_col(y), value="-")
                    _apply_style(ws.cell(row=row, column=yr_col(y)), "data")
                else:
                    # =YTM_Base + StepUp * (year-1)
                    formula = f"={amt_cell_base}+{amt_cell_stepup}*{y - 1}"
                    ws.cell(row=row, column=yr_col(y), value=formula)
                    _apply_style(ws.cell(row=row, column=yr_col(y)), "formula", NUM_FMT_PCT)
        row += 1

        # Accumulated factor: (1+ytm)^year
        ytm_accum_rows[ti] = row
        ws.cell(row=row, column=label_col, value=f"  {tname} 누적 Factor")
        _apply_style(ws.cell(row=row, column=label_col), "label")
        ws.cell(row=row, column=unit_col, value="x")
        _apply_style(ws.cell(row=row, column=unit_col), "data")

        for y in year_cols:
            if y == 0:
                ws.cell(row=row, column=yr_col(y), value=1.0)
                _apply_style(ws.cell(row=row, column=yr_col(y)), "data")
            else:
                ytm_cell = yr_cell(ytm_schedule_rows[ti], y)
                prev_factor = yr_cell(row, y - 1)
                formula = f"={prev_factor}*(1+{ytm_cell})"
                ws.cell(row=row, column=yr_col(y), value=formula)
                _apply_style(ws.cell(row=row, column=yr_col(y)), "formula", "0.0000")
        row += 1

    row += 1  # blank

    # ══════════════════════════════════════════════════
    # SECTION: Cash-out (투자 집행)
    # ══════════════════════════════════════════════════
    ws.cell(row=row, column=label_col, value="▶ Cash-out (투자 집행)")
    _apply_style(ws.cell(row=row, column=label_col), "section")
    for y in year_cols:
        _apply_style(ws.cell(row=row, column=yr_col(y)), "section")
    _apply_style(ws.cell(row=row, column=unit_col), "section")
    row += 1

    cashout_rows: Dict[int, int] = {}
    for ti, tr in enumerate(tranches):
        tname = tr.get("name", f"Tranche {chr(65 + ti)}")
        cashout_rows[ti] = row
        ws.cell(row=row, column=label_col, value=f"  {tname}")
        _apply_style(ws.cell(row=row, column=label_col), "label")
        ws.cell(row=row, column=unit_col, value="억원")
        _apply_style(ws.cell(row=row, column=unit_col), "data")

        amt_ref = cell_ref(input_rows[f"t{ti}_amount"], yr_col(0))
        for y in year_cols:
            if y == 0:
                formula = f"=-{amt_ref}"
                ws.cell(row=row, column=yr_col(y), value=formula)
                _apply_style(ws.cell(row=row, column=yr_col(y)), "formula", NUM_FMT_AMOUNT)
            else:
                ws.cell(row=row, column=yr_col(y), value=0)
                _apply_style(ws.cell(row=row, column=yr_col(y)), "data", NUM_FMT_AMOUNT)
        row += 1

    # Total cash-out
    total_cashout_row = row
    ws.cell(row=row, column=label_col, value="  Total Cash-out")
    _apply_style(ws.cell(row=row, column=label_col), "label_bold")
    ws.cell(row=row, column=unit_col, value="억원")
    _apply_style(ws.cell(row=row, column=unit_col), "subtotal")
    for y in year_cols:
        refs = [yr_cell(cashout_rows[ti], y) for ti in range(len(tranches))]
        formula = "=" + "+".join(refs)
        ws.cell(row=row, column=yr_col(y), value=formula)
        _apply_style(ws.cell(row=row, column=yr_col(y)), "subtotal", NUM_FMT_AMOUNT)
    row += 2

    # ══════════════════════════════════════════════════
    # SECTION: Tranche Cash Flows
    # ══════════════════════════════════════════════════
    tranche_cf_rows: Dict[int, Dict[str, int]] = {}  # ti -> {label: row}

    for ti, tr in enumerate(tranches):
        tname = tr.get("name", f"Tranche {chr(65 + ti)}")
        ttype = tr.get("type", "CB")
        maturity = _safe_int(tr.get("maturity_years", max_maturity))
        coupon = _safe_float(tr.get("coupon_rate", 0))
        div_rate = _safe_float(tr.get("dividend_rate", 0))
        call_year = _safe_int(tr.get("call_year", 0))

        tranche_cf_rows[ti] = {}

        ws.cell(row=row, column=label_col, value=f"▶ {tname} Cash Flow ({ttype})")
        _apply_style(ws.cell(row=row, column=label_col), "section")
        for y in year_cols:
            _apply_style(ws.cell(row=row, column=yr_col(y)), "section")
        _apply_style(ws.cell(row=row, column=unit_col), "section")
        row += 1

        amt_ref = cell_ref(input_rows[f"t{ti}_amount"], yr_col(0))
        coupon_ref = cell_ref(input_rows[f"t{ti}_coupon"], yr_col(0))

        # Coupon income
        tranche_cf_rows[ti]["coupon"] = row
        ws.cell(row=row, column=label_col, value="  Coupon 수취")
        _apply_style(ws.cell(row=row, column=label_col), "label")
        ws.cell(row=row, column=unit_col, value="억원")
        _apply_style(ws.cell(row=row, column=unit_col), "data")
        for y in year_cols:
            if y == 0:
                ws.cell(row=row, column=yr_col(y), value=0)
                _apply_style(ws.cell(row=row, column=yr_col(y)), "data", NUM_FMT_AMOUNT)
            elif y <= maturity:
                formula = f"={amt_ref}*{coupon_ref}"
                ws.cell(row=row, column=yr_col(y), value=formula)
                _apply_style(ws.cell(row=row, column=yr_col(y)), "formula", NUM_FMT_AMOUNT)
            else:
                ws.cell(row=row, column=yr_col(y), value=0)
                _apply_style(ws.cell(row=row, column=yr_col(y)), "data", NUM_FMT_AMOUNT)
        row += 1

        # Dividend (for CPS/PS/RCPS)
        if ttype in ("CPS", "PS", "RCPS") and f"t{ti}_div_rate" in input_rows:
            tranche_cf_rows[ti]["dividend"] = row
            div_ref = cell_ref(input_rows[f"t{ti}_div_rate"], yr_col(0))
            ws.cell(row=row, column=label_col, value="  우선배당")
            _apply_style(ws.cell(row=row, column=label_col), "label")
            ws.cell(row=row, column=unit_col, value="억원")
            _apply_style(ws.cell(row=row, column=unit_col), "data")
            for y in year_cols:
                if y == 0:
                    ws.cell(row=row, column=yr_col(y), value=0)
                    _apply_style(ws.cell(row=row, column=yr_col(y)), "data", NUM_FMT_AMOUNT)
                elif y <= maturity:
                    formula = f"={amt_ref}*{div_ref}"
                    ws.cell(row=row, column=yr_col(y), value=formula)
                    _apply_style(ws.cell(row=row, column=yr_col(y)), "formula", NUM_FMT_AMOUNT)
                else:
                    ws.cell(row=row, column=yr_col(y), value=0)
                    _apply_style(ws.cell(row=row, column=yr_col(y)), "data", NUM_FMT_AMOUNT)
            row += 1

        # YTM accrual (paper gain, shown for reference)
        tranche_cf_rows[ti]["ytm_accrual"] = row
        ws.cell(row=row, column=label_col, value="  YTM 정산액 (Put/만기)")
        _apply_style(ws.cell(row=row, column=label_col), "label")
        ws.cell(row=row, column=unit_col, value="억원")
        _apply_style(ws.cell(row=row, column=unit_col), "data")

        factor_row = ytm_accum_rows.get(ti)
        for y in year_cols:
            if y == 0:
                ws.cell(row=row, column=yr_col(y), value=0)
                _apply_style(ws.cell(row=row, column=yr_col(y)), "data", NUM_FMT_AMOUNT)
            elif y == maturity:
                # At maturity: Amount * (Factor - 1) = YTM premium
                if factor_row:
                    factor_cell = yr_cell(factor_row, y)
                    formula = f"={amt_ref}*({factor_cell}-1)"
                else:
                    formula = f"=0"
                ws.cell(row=row, column=yr_col(y), value=formula)
                _apply_style(ws.cell(row=row, column=yr_col(y)), "formula", NUM_FMT_AMOUNT)
            else:
                ws.cell(row=row, column=yr_col(y), value=0)
                _apply_style(ws.cell(row=row, column=yr_col(y)), "data", NUM_FMT_AMOUNT)
        row += 1

        # Maturity repayment (원금 상환)
        tranche_cf_rows[ti]["maturity_repay"] = row
        ws.cell(row=row, column=label_col, value="  만기 상환 (원금)")
        _apply_style(ws.cell(row=row, column=label_col), "label")
        ws.cell(row=row, column=unit_col, value="억원")
        _apply_style(ws.cell(row=row, column=unit_col), "data")
        for y in year_cols:
            if y == maturity:
                formula = f"={amt_ref}"
                ws.cell(row=row, column=yr_col(y), value=formula)
                _apply_style(ws.cell(row=row, column=yr_col(y)), "formula", NUM_FMT_AMOUNT)
            else:
                ws.cell(row=row, column=yr_col(y), value=0)
                _apply_style(ws.cell(row=row, column=yr_col(y)), "data", NUM_FMT_AMOUNT)
        row += 1

        # Tranche subtotal
        tranche_cf_rows[ti]["subtotal"] = row
        ws.cell(row=row, column=label_col, value=f"  {tname} Net CF")
        _apply_style(ws.cell(row=row, column=label_col), "label_bold")
        ws.cell(row=row, column=unit_col, value="억원")
        _apply_style(ws.cell(row=row, column=unit_col), "subtotal")

        cf_keys = ["coupon", "dividend", "ytm_accrual", "maturity_repay"]
        cf_line_rows = [tranche_cf_rows[ti][k] for k in cf_keys if k in tranche_cf_rows[ti]]

        for y in year_cols:
            if y == 0:
                ws.cell(row=row, column=yr_col(y), value=0)
                _apply_style(ws.cell(row=row, column=yr_col(y)), "subtotal", NUM_FMT_AMOUNT)
            else:
                refs = [yr_cell(r, y) for r in cf_line_rows]
                formula = "=" + "+".join(refs)
                ws.cell(row=row, column=yr_col(y), value=formula)
                _apply_style(ws.cell(row=row, column=yr_col(y)), "subtotal", NUM_FMT_AMOUNT)
        row += 2

    # ══════════════════════════════════════════════════
    # SECTION: Total Net Cash Flow
    # ══════════════════════════════════════════════════
    ws.cell(row=row, column=label_col, value="▶ Total Net Cash Flow")
    _apply_style(ws.cell(row=row, column=label_col), "section")
    for y in year_cols:
        _apply_style(ws.cell(row=row, column=yr_col(y)), "section")
    _apply_style(ws.cell(row=row, column=unit_col), "section")
    row += 1

    total_net_cf_row = row
    ws.cell(row=row, column=label_col, value="  Total Net CF")
    _apply_style(ws.cell(row=row, column=label_col), "label_bold")
    ws.cell(row=row, column=unit_col, value="억원")
    _apply_style(ws.cell(row=row, column=unit_col), "subtotal")

    for y in year_cols:
        refs = []
        refs.append(yr_cell(total_cashout_row, y))
        for ti in range(len(tranches)):
            if "subtotal" in tranche_cf_rows.get(ti, {}):
                refs.append(yr_cell(tranche_cf_rows[ti]["subtotal"], y))
        formula = "=" + "+".join(refs)
        ws.cell(row=row, column=yr_col(y), value=formula)
        _apply_style(ws.cell(row=row, column=yr_col(y)), "subtotal", NUM_FMT_AMOUNT)
    row += 1

    # IRR (base case)
    base_irr_row = row
    ws.cell(row=row, column=label_col, value="  IRR (Base)")
    _apply_style(ws.cell(row=row, column=label_col), "label_bold")
    ws.cell(row=row, column=unit_col, value="%")
    _apply_style(ws.cell(row=row, column=unit_col), "irr")
    range_start = yr_cell(total_net_cf_row, 0)
    range_end = yr_cell(total_net_cf_row, n_years)
    formula = f'=IFERROR(IRR({range_start}:{range_end}),"-")'
    ws.cell(row=row, column=yr_col(0), value=formula)
    _apply_style(ws.cell(row=row, column=yr_col(0)), "irr", NUM_FMT_PCT)
    row += 2

    # ══════════════════════════════════════════════════
    # SECTION: Exit Scenarios
    # ══════════════════════════════════════════════════
    if exit_scenarios:
        ws.cell(row=row, column=label_col, value="▶ Exit Scenario Analysis")
        _apply_style(ws.cell(row=row, column=label_col), "section")
        for y in year_cols:
            _apply_style(ws.cell(row=row, column=yr_col(y)), "section")
        _apply_style(ws.cell(row=row, column=unit_col), "section")
        row += 1

        for si, scenario in enumerate(exit_scenarios):
            sname = scenario.get("name", f"Scenario {si + 1}")
            exit_year = _safe_int(scenario.get("exit_year", max_maturity))
            multiple = _safe_float(scenario.get("multiple", 1.0))

            if exit_year > n_years:
                exit_year = n_years

            # Scenario CF row
            scen_cf_row = row
            ws.cell(row=row, column=label_col, value=f"  {sname}")
            _apply_style(ws.cell(row=row, column=label_col), "label")
            ws.cell(row=row, column=unit_col, value="억원")
            _apply_style(ws.cell(row=row, column=unit_col), "data")

            # Total investment amount cell (sum of all tranche amounts)
            total_amt_refs = [cell_ref(input_rows[f"t{ti}_amount"], yr_col(0))
                              for ti in range(len(tranches))]
            total_amt_formula = "+".join(total_amt_refs)

            for y in year_cols:
                if y == 0:
                    # Cash out = negative total
                    formula = f"=-({total_amt_formula})"
                    ws.cell(row=row, column=yr_col(y), value=formula)
                    _apply_style(ws.cell(row=row, column=yr_col(y)), "formula", NUM_FMT_AMOUNT)
                elif y == exit_year:
                    # Exit proceeds = total * multiple + accumulated coupons/dividends
                    formula = f"=({total_amt_formula})*{multiple}"
                    ws.cell(row=row, column=yr_col(y), value=formula)
                    _apply_style(ws.cell(row=row, column=yr_col(y)), "formula", NUM_FMT_AMOUNT)
                elif y < exit_year:
                    # Intermediate coupon/dividend income from total net CF
                    ref = yr_cell(total_net_cf_row, y)
                    # Only inflows (coupon/dividend), not the outflow
                    coupon_refs = []
                    for ti in range(len(tranches)):
                        if "coupon" in tranche_cf_rows.get(ti, {}):
                            coupon_refs.append(yr_cell(tranche_cf_rows[ti]["coupon"], y))
                        if "dividend" in tranche_cf_rows.get(ti, {}):
                            coupon_refs.append(yr_cell(tranche_cf_rows[ti]["dividend"], y))
                    if coupon_refs:
                        formula = "=" + "+".join(coupon_refs)
                    else:
                        formula = "=0"
                    ws.cell(row=row, column=yr_col(y), value=formula)
                    _apply_style(ws.cell(row=row, column=yr_col(y)), "formula", NUM_FMT_AMOUNT)
                else:
                    ws.cell(row=row, column=yr_col(y), value=0)
                    _apply_style(ws.cell(row=row, column=yr_col(y)), "data", NUM_FMT_AMOUNT)
            row += 1

            # Scenario IRR
            ws.cell(row=row, column=label_col, value=f"    → IRR")
            _apply_style(ws.cell(row=row, column=label_col), "label")
            ws.cell(row=row, column=unit_col, value="%")
            _apply_style(ws.cell(row=row, column=unit_col), "irr")
            r_start = yr_cell(scen_cf_row, 0)
            r_end = yr_cell(scen_cf_row, n_years)
            formula = f'=IFERROR(IRR({r_start}:{r_end}),"-")'
            ws.cell(row=row, column=yr_col(0), value=formula)
            _apply_style(ws.cell(row=row, column=yr_col(0)), "irr", NUM_FMT_PCT)

            # MoM (Money on Money)
            ws.cell(row=row, column=yr_col(1), value="MoM:")
            _apply_style(ws.cell(row=row, column=yr_col(1)), "label")
            exit_cell = yr_cell(scen_cf_row, exit_year)
            invest_cell = yr_cell(scen_cf_row, 0)
            formula = f'=IFERROR(-{exit_cell}/{invest_cell},"-")'
            ws.cell(row=row, column=yr_col(2), value=formula)
            _apply_style(ws.cell(row=row, column=yr_col(2)), "irr", "0.00x")
            row += 1

        row += 1

    # ══════════════════════════════════════════════════
    # SECTION: Projections (if available)
    # ══════════════════════════════════════════════════
    if projections and projections.get("years"):
        ws.cell(row=row, column=label_col, value="▶ Financial Projections")
        _apply_style(ws.cell(row=row, column=label_col), "section")
        for y in year_cols:
            _apply_style(ws.cell(row=row, column=yr_col(y)), "section")
        _apply_style(ws.cell(row=row, column=unit_col), "section")
        row += 1

        proj_years = projections["years"]

        # Map projection years to model year columns
        for metric_key, metric_label in [
            ("revenue", "매출"),
            ("ebitda", "EBITDA"),
            ("net_income", "순이익"),
        ]:
            values = projections.get(metric_key)
            if not values:
                continue
            ws.cell(row=row, column=label_col, value=f"  {metric_label}")
            _apply_style(ws.cell(row=row, column=label_col), "label")
            ws.cell(row=row, column=unit_col, value="억원")
            _apply_style(ws.cell(row=row, column=unit_col), "data")

            for y in year_cols:
                if y < len(values):
                    ws.cell(row=row, column=yr_col(y), value=_safe_float(values[y]))
                    _apply_style(ws.cell(row=row, column=yr_col(y)), "input", NUM_FMT_AMOUNT)
                else:
                    ws.cell(row=row, column=yr_col(y), value=0)
                    _apply_style(ws.cell(row=row, column=yr_col(y)), "data", NUM_FMT_AMOUNT)
            row += 1

        # Projection year labels
        ws.cell(row=row, column=label_col, value="  (연도)")
        _apply_style(ws.cell(row=row, column=label_col), "label")
        for y in year_cols:
            if y < len(proj_years):
                ws.cell(row=row, column=yr_col(y), value=proj_years[y])
                _apply_style(ws.cell(row=row, column=yr_col(y)), "data")
        row += 1

    # ══════════════════════════════════════════════════
    # SHEET 2: Covenants
    # ══════════════════════════════════════════════════
    if covenants:
        ws2 = wb.create_sheet("Covenants")
        ws2.column_dimensions["A"].width = 6
        ws2.column_dimensions["B"].width = 16
        ws2.column_dimensions["C"].width = 60

        r = 1
        ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        ws2.cell(row=r, column=1, value=f"{company} — Covenants & Governance")
        _apply_style(ws2.cell(row=r, column=1), "title")
        ws2.row_dimensions[r].height = 28
        r += 2

        # Headers
        for ci, header in enumerate(["No.", "구분", "내용"], 1):
            c = ws2.cell(row=r, column=ci, value=header)
            _apply_style(c, "header")
        r += 1

        for ci, cov in enumerate(covenants, 1):
            ws2.cell(row=r, column=1, value=ci)
            _apply_style(ws2.cell(row=r, column=1), "data")
            cat = cov.get("category", "기타")
            ws2.cell(row=r, column=2, value=cat)
            _apply_style(ws2.cell(row=r, column=2), "data")
            ws2.cell(row=r, column=3, value=cov.get("description", ""))
            _apply_style(ws2.cell(row=r, column=3), "data")
            ws2.cell(row=r, column=3).alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True)
            r += 1

    # ── Freeze panes ──
    ws.freeze_panes = f"{get_column_letter(first_year_col)}{header_row + 1}"

    # ── Save ──
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ── Combined: Extract + Build ──────────────────────────

def generate_excel_model(
    api_key: str,
    model_name: str,
    project_name: str,
    owner_id: int | None = None,
) -> dict:
    """Full pipeline: extract structure via AI, then build Excel model.
    Returns {"structure": dict, "excel_bytes": bytes (base64), "filename": str}.
    """
    import base64

    structure = extract_deal_structure(api_key, model_name, project_name, owner_id)
    excel_bytes = build_excel_model(structure)

    company = structure.get("company_name", project_name)
    filename = f"{company}_CashFlow_Model.xlsx"

    return {
        "structure": structure,
        "excel_b64": base64.b64encode(excel_bytes).decode("ascii"),
        "filename": filename,
    }
